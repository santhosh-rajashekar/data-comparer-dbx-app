"""File Service — Upload and parse Excel/CSV files.

Replaces the SheetJS (xlsx.full.min.js) client-side parsing
from the original HTML app with server-side openpyxl + pandas.
"""

import os
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from typing import Optional


class FileService:
    """Handles file upload, parsing, and storage."""

    def __init__(self, volume_path: str = "/Volumes/data_mesh_hub/rdm/uploads"):
        self.volume_path = volume_path

    def process_upload(self, file, source: str, session_id: Optional[str] = None) -> dict:
        """Process an uploaded file (Excel or CSV).

        Args:
            file: Flask FileStorage object.
            source: Source identifier ('COA', 'FAQ', 'DataPool').
            session_id: Optional session ID for file organization.

        Returns:
            Dict with filename, headers, row_count, sheets (for Excel).
        """
        filename = file.filename
        content = file.read()

        if filename.endswith((".xlsx", ".xls")):
            return self._parse_excel(content, filename, source)
        elif filename.endswith(".xlsb"):
            return self._parse_xlsb(content, filename, source)
        elif filename.endswith(".csv"):
            return self._parse_csv(content, filename, source)
        else:
            raise ValueError(f"Unsupported file format: {filename}")

    def _parse_excel(self, content: bytes, filename: str, source: str) -> dict:
        """Parse Excel file using openpyxl.

        Handles:
        - Multi-sheet workbooks
        - Yellow-row filtering (SKB mode)
        - Strikethrough detection
        """
        wb = load_workbook(io.BytesIO(content), data_only=True)
        sheets = wb.sheetnames

        # Default to first sheet
        ws = wb.active
        rows_data = []
        headers = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            if row_idx == 1:
                headers = [str(cell.value or "").strip() for cell in row]
                continue

            # Skip yellow-highlighted rows (COA filtering)
            if source == "COA" and self._is_yellow_row(row):
                continue

            # Skip strikethrough rows
            if self._is_strike_row(row):
                continue

            row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
            if any(v.strip() for v in row_values):  # Skip fully empty rows
                rows_data.append(row_values)

        return {
            "filename": filename,
            "source": source,
            "headers": headers,
            "rows": rows_data,
            "row_count": len(rows_data),
            "sheets": sheets,
            "active_sheet": ws.title,
        }

    def _parse_csv(self, content: bytes, filename: str, source: str) -> dict:
        """Parse CSV file using pandas."""
        # Try different encodings
        for encoding in ["utf-8", "latin-1", "cp1252"]:
            try:
                df = pd.read_csv(io.BytesIO(content), encoding=encoding, dtype=str)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError(f"Unable to decode {filename} with common encodings")

        df = df.fillna("")
        headers = df.columns.tolist()
        rows_data = df.values.tolist()

        return {
            "filename": filename,
            "source": source,
            "headers": headers,
            "rows": rows_data,
            "row_count": len(rows_data),
            "sheets": [],
            "active_sheet": None,
        }

    def _parse_xlsb(self, content: bytes, filename: str, source: str) -> dict:
        """Parse Excel Binary Workbook (.xlsb) using pandas + pyxlsb.

        openpyxl does not support .xlsb format, so we use pandas with
        the pyxlsb engine. Note: style-based filtering (yellow rows,
        strikethrough) is not available for .xlsb files.
        """
        import sys
        print(f"[FileService] Parsing .xlsb file: {filename} ({len(content)} bytes)", flush=True)

        buffer = io.BytesIO(content)
        xls = pd.ExcelFile(buffer, engine="pyxlsb")
        sheets = xls.sheet_names
        print(f"[FileService] Sheets found: {sheets}", flush=True)

        # Read first sheet with minimal memory: only keep string dtype
        df = pd.read_excel(xls, sheet_name=0, dtype=str)
        df = df.fillna("")

        # Drop fully empty rows efficiently
        mask = df.apply(lambda row: row.str.strip().any(), axis=1)
        df = df[mask].reset_index(drop=True)

        headers = df.columns.tolist()
        row_count = len(df)
        print(f"[FileService] Parsed {row_count} rows, {len(headers)} columns", flush=True)

        # Store rows as list (needed for diff engine later)
        rows_data = df.values.tolist()

        # Free the DataFrame to reduce memory
        del df
        del buffer

        return {
            "filename": filename,
            "source": source,
            "headers": headers,
            "rows": rows_data,
            "row_count": row_count,
            "sheets": sheets,
            "active_sheet": sheets[0] if sheets else None,
        }

    @staticmethod
    def _is_yellow_row(row) -> bool:
        """Check if a row has yellow background (COA exclusion marker)."""
        for cell in row[:5]:  # Check first 5 cells
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                rgb = str(cell.fill.fgColor.rgb)
                if len(rgb) >= 6:
                    # Extract RGB and check if yellowish
                    try:
                        hex_color = rgb[-6:]
                        r = int(hex_color[0:2], 16)
                        g = int(hex_color[2:4], 16)
                        b = int(hex_color[4:6], 16)
                        if r > 200 and g > 200 and b < 100:
                            return True
                    except (ValueError, IndexError):
                        pass
        return False

    @staticmethod
    def _is_strike_row(row) -> bool:
        """Check if a row has strikethrough formatting."""
        strike_count = 0
        for cell in row[:5]:
            if cell.font and cell.font.strike:
                strike_count += 1
        return strike_count >= 3  # Majority of checked cells are struck through
